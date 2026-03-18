"""
FORMATEADOR DE EXCEL
Utilidades para exportar a Excel, CSV y JSON con formato profesional
"""

import pandas as pd
from io import BytesIO
import json
from typing import Any
import logging
from openpyxl.styles import Font, PatternFill, Alignment, numbers

logger = logging.getLogger(__name__)


def _normalizar_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normaliza un DataFrame para exportación:
    - Columnas en MAYÚSCULAS
    - Texto en MAYÚSCULAS
    - Fechas en formato DD/MM/YYYY (sin T, sin decimales, sin hora si es medianoche)
    """
    df = df.copy()

    # Renombrar columnas a MAYÚSCULAS
    df.columns = [str(col).upper() for col in df.columns]

    def _limpiar_fecha(val_str: str) -> str:
        """Convierte cualquier formato de fecha/datetime a DD/MM/YYYY o DD/MM/YYYY HH:MM"""
        try:
            # Soporta: "2026-03-26T00:00:00", "2026-03-26 00:00:00.000000", "2026-03-26"
            dt = pd.to_datetime(val_str)
            if dt.hour == 0 and dt.minute == 0 and dt.second == 0:
                return dt.strftime("%d/%m/%Y")
            return dt.strftime("%d/%m/%Y %H:%M")
        except Exception:
            return str(val_str).upper()

    def _limpiar_valor(val) -> str:
        if pd.isna(val) or val is None:
            return val
        val_str = str(val).strip()
        # Detectar fechas: empieza con 4 dígitos-mes-día (ISO) o contiene T entre fecha y hora
        es_fecha = (
            (len(val_str) >= 10 and val_str[4:5] == '-' and val_str[7:8] == '-') or
            ('T' in val_str and val_str[4:5] == '-')
        )
        if es_fecha:
            return _limpiar_fecha(val_str)
        return val_str.upper()

    # Procesar cada columna
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            # Columnas datetime nativas de pandas
            df[col] = df[col].apply(
                lambda dt: dt.strftime("%d/%m/%Y") if pd.notna(dt) and dt.hour == 0 and dt.minute == 0
                else (dt.strftime("%d/%m/%Y %H:%M") if pd.notna(dt) else "")
            )
        elif df[col].dtype == object:
            df[col] = df[col].apply(_limpiar_valor)

    return df



class ExcelFormatter:
    """Formateador para exportaciones de datos"""
    
    @staticmethod
    def crear_excel(df: pd.DataFrame, titulo: str = "Reporte") -> bytes:
        """
        Crea archivo Excel con formato profesional
        - Texto en MAYÚSCULAS
        - Fechas limpias
        - Encabezados con estilo
        """
        try:
            # Normalizar datos
            df = _normalizar_df(df)
            
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Datos', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Datos']
                
                # Estilo de encabezados
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(patternType="solid", fgColor="1F4E78")
                header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                
                # Formato de celdas de datos
                data_align = Alignment(vertical="center")
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        cell.alignment = data_align
                
                # Ajustar ancho de columnas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 3, 50)
                    worksheet.column_dimensions[column_letter].width = max(adjusted_width, 12)
                
                # Congelar primera fila (encabezados)
                worksheet.freeze_panes = 'A2'
            
            output.seek(0)
            return output.read()
        
        except Exception as e:
            logger.error(f"Error creando Excel: {str(e)}")
            raise
    
    @staticmethod
    def crear_csv(df: pd.DataFrame) -> bytes:
        """
        Crea archivo CSV con texto en MAYÚSCULAS
        """
        try:
            df = _normalizar_df(df)
            output = BytesIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)
            return output.read()
        
        except Exception as e:
            logger.error(f"Error creando CSV: {str(e)}")
            raise
    
    @staticmethod
    def crear_json(df: pd.DataFrame) -> bytes:
        """
        Crea archivo JSON con texto en MAYÚSCULAS
        """
        try:
            df = _normalizar_df(df)
            datos = df.to_dict(orient='records')
            json_str = json.dumps(datos, ensure_ascii=False, indent=2)
            return json_str.encode('utf-8')
        
        except Exception as e:
            logger.error(f"Error creando JSON: {str(e)}")
            raise
